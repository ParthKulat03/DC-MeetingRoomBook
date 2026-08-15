from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook

from app.core.exceptions import ConflictException
from app.repositories.role_repository import RoleRepository
from app.repositories.user_repository import UserRepository
from app.schemas.user import UserCreateRequest
from app.services.users.user_service import UserService


class UserImportService:

    def __init__(
        self,
        user_repository: UserRepository,
        role_repository: RoleRepository,
        user_service: UserService,
    ):
        self.user_repository = user_repository
        self.role_repository = role_repository
        self.user_service = user_service

    async def import_file(
        self,
        file: BinaryIO,
    ) -> dict:

        contents = file.read()

        workbook = load_workbook(
            filename=BytesIO(contents),
            read_only=True,
        )

        worksheet = workbook.active

        total = 0
        imported = 0
        updated = 0
        invalid = 0
        duplicate = 0
        failed = 0

        for row in worksheet.iter_rows(
            values_only=True
        ):

            if not row or not row[0]:
                continue

            total += 1

            try:

                employee_id = str(row[0]).strip()
                name = str(row[1]).strip()
                email = str(row[2]).strip()

                designation = (
                    str(row[3]).strip()
                    if len(row) > 3
                    and row[3] is not None
                    else None
                )

                existing = (
                    await self.user_repository
                    .get_by_employee_id(
                        employee_id
                    )
                )

                if existing:
                    updated += 1
                    continue

                await self.user_service.create_user(
                    UserCreateRequest(
                        employee_id=employee_id,
                        name=name,
                        email=email,
                        designation=designation,
                        role="EMPLOYEE",
                    )
                )

                imported += 1

            except ValueError:
                invalid += 1

            except ConflictException:
                duplicate += 1

            except Exception:
                failed += 1

        workbook.close()

        return {
            "total": total,
            "imported": imported,
            "updated": updated,
            "invalid": invalid,
            "duplicate": duplicate,
            "failed": failed,
            "errorReportUrl": None,
        }